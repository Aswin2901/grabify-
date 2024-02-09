import json
from operator import itemgetter
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponseNotAllowed, HttpResponseRedirect,HttpResponse
from django.shortcuts import render,redirect , get_object_or_404
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.contrib.auth import authenticate, login
from grabify.models import CustomUser as grabify_user , OrderDetails , OrderItems
from .models import Product , Category , Variant , Offer
from .forms import ProductFilterForm , UserFilterForm , OrderFilterForm ,OfferForm
from datetime import timedelta
from django.db.models import Sum , Q
from django.template.loader import get_template
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views import View
from xhtml2pdf import pisa
import openpyxl
from django.http import JsonResponse
from grabify.views import logout as custom_logout




# Create your views here.

def admin_login(request):
    return render(request, 'admin_login.html')

def dashbord(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    product_list = Product.objects.all()
    category_list = Category.objects.all()
    user_list = grabify_user.objects.all()
    order_list = OrderDetails.objects.all()
    pending_orders_count = OrderDetails.objects.exclude(Q(status='Delivered') | Q(status='Cancelled')).count()
    return render(request , 'dashbord.html',{'pending_orders_count' : pending_orders_count , 'product_list' : product_list ,'order_list' : order_list , 'category_list' : category_list , 'user_list' : user_list})


def user_list(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    filter_form = UserFilterForm(request.GET)
    search_query = request.GET.get('search_query')
    
    if search_query:
        users = grabify_user.objects.filter(fullname__icontains=search_query) | grabify_user.objects.filter(email__icontains=search_query)
    elif filter_form.is_valid():
        status = filter_form.cleaned_data.get('status')
        if status and status != 'all':
            users = grabify_user.objects.filter(status=status)
        else:
            users = grabify_user.objects.all()
    else:
        users = grabify_user.objects.all()

    context = {'admin_user': users, 'filter_form': filter_form}
    return render(request, "admin_user.html", context)

def block_user(request, user_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    user = grabify_user.objects.get(pk=user_id)
    user.status = 'blocked'
    user.save()
    return redirect('custom_admin:user_list')

def unblock_user(request, user_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    user = grabify_user.objects.get(pk=user_id)
    user.status = 'active'
    user.save()
    return redirect('custom_admin:user_list')
    

def add_product(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    categories = Category.objects.all()
    
    if request.method == 'POST':
        name = request.POST['name']
        image = request.FILES['image']
        description = request.POST['description']
        price = request.POST['price']
        category_id = request.POST['category']
        status = request.POST.get('status', 'active')
        category = Category.objects.get(pk=category_id)
        variant_images = request.FILES.getlist('variants', [])
        # Create a new Product instance and save it
        product = Product.objects.create(
            name=name,
            image=image,
            description=description,
            price=price,
            category=category,
            status=status
        )
        
        for variant_image in variant_images:
            Variant.objects.create(product=product, image=variant_image)

        
        product.save()
        return redirect('custom_admin:add_variants', product_id=product.id)  # Redirect to product list or another appropriate page
    return render(request, 'add_product.html', {'categories': categories})

def product_list(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    filter_form = ProductFilterForm(request.GET)
    products_with_variant_quantities = []

    search_query = request.GET.get('search_query')

    if search_query:
        products = Product.objects.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query)).order_by('-created_at')
    elif filter_form.is_valid():
        status = filter_form.cleaned_data.get('status')
        products = Product.objects.filter(status=status).order_by('-created_at') if status and status != 'all' else Product.objects.all().order_by('-created_at')
    else:
        products = Product.objects.all().order_by('-created_at')

    for product in products:
        variant_quantities = Variant.objects.filter(product=product).values('size').annotate(quantity=Sum('quantity'))
        variant_quantities_dict = {v['size']: v['quantity'] for v in variant_quantities}
        products_with_variant_quantities.append({'product': product, 'variant_quantities': variant_quantities_dict})

        # Update status based on quantity
        product_quantity = product.quantity
        if product_quantity:
            if product.status != 'inactive':
                if product_quantity <= 0:
                    product.status = 'outofstock'
                else:
                    product.status = 'active'
                product.save()
            
        else:
            product.status = 'outofstock'
            product.save()

    return render(request, "products.html", {'products_with_variant_quantities': products_with_variant_quantities, 'filter_form': filter_form})

def edit_product(request, product_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    product = get_object_or_404(Product, id=product_id)
    categories = Category.objects.all()
    variant_images = Variant.objects.filter(product=product)

    if request.method == 'POST':
        # Process form submission
        product.name = request.POST.get('name')
        product.image = request.FILES.get('image', product.image)
        product.category_id = request.POST.get('category')
        
        product.price = float(request.POST.get('price', 0))

        product.description = request.POST.get('description')
        product.status = request.POST.get('status')
        product.save()

        # Process variant images
        variant_images = request.FILES.getlist('variants', [])
        for variant_image in variant_images:
            Variant.objects.update_or_create(product=product, image=variant_image)

        # Process removed images
        removed_images_json = request.POST.get('removed_images', '[]')
        removed_images = []
        if removed_images_json:
            removed_images = json.loads(removed_images_json)
        for removed_image in removed_images:
            image_id = removed_image.get('id')
            try:
                image = Variant.objects.get(id=image_id)
                image.delete()
            except Variant.DoesNotExist:
                pass
                    

        return redirect('custom_admin:product_list')

    edit_mode = request.GET.get('edit_mode', False)
    return render(request, 'add_product.html', {'edit_mode': edit_mode, 'variant_images': variant_images, 'categories': categories, 'product': product})

    
def deactivate_product(request, product_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    if request.method == 'POST':
        product = get_object_or_404(Product, pk=product_id)
        product.status = 'inactive'
        product.save()
        print(product.status)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
def activate_product(request, product_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    product = get_object_or_404(Product, pk=product_id)
    product.status = 'active'
    product.save()
    return redirect('custom_admin:product_list')  


def category_list(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    context = {'categories': Category.objects.all()}
    return render(request , "category.html",context)

def add_category(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        
        category = Category.objects.create(
            name=name,
            description=description,
        )
        
        category.save()
        return redirect('custom_admin:category_list')  # Redirect to category list or another appropriate page
    return render(request, 'add_category.html', {'edit_mode': False})

def edit_category(request, category_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    category = get_object_or_404(Category, id=category_id)
    
    if request.method == 'POST':
        # Process form submission
        category.name = request.POST.get('name')
        category.description = request.POST.get('description')
        category.save()
        return redirect('custom_admin:category_list')
    
    edit_mode = request.GET.get('edit_mode', False)
    return render(request, 'add_category.html', {'edit_mode': edit_mode, 'category': category})

def activate_category(request, category_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    category = get_object_or_404(Category, id=category_id)
    category.status = 'active'
    category.save()
    return redirect('custom_admin:category_list')

def deactivate_category(request, category_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    category = get_object_or_404(Category, id=category_id)
    category.status = 'inactive'
    category.save()
    return redirect('custom_admin:category_list')


def order_management(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    filter_form = OrderFilterForm(request.GET)
    if filter_form.is_valid():
        status = filter_form.cleaned_data.get('status')
        print(status)
        if status and status != 'all':
            orders = OrderDetails.objects.filter(status=status)
        else:
            orders = OrderDetails.objects.all()
    else:
        orders = OrderDetails.objects.all()

    # Create a list to store the data for each order
    order_data = []
    for order in orders:
        user_name = order.user.fullname  
        product_name = ", ".join(item.product.name for item in OrderItems.objects.filter(order_id=order))
        order_date = order.created_at.strftime("%d %b %Y")
        arriving_date = order.created_at + timedelta(days=5)
        status = order.status

        # Append data for each order to the list
        order_data.append({
            'order_id': order.id,
            'user_name': user_name,
            'product_name': product_name,
            'order_date': order_date,
            'arriving_date': arriving_date,
            'status': status,
        })

    order_data = sorted(order_data, key=itemgetter('order_id'), reverse=True)
    return render(request, 'order_management.html', {'order_data': order_data, 'filter_form': filter_form})

def change_order_status(request , order_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    order = get_object_or_404(OrderDetails , id = order_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status == 'Accept':
            order.status = 'Accepted'
            order.save()
        elif new_status == 'Cancel':
            order.status = 'Cancelled'
            order.save()
        elif new_status == 'placed':
            order.status = 'Placed'
            order.save()    
        elif new_status == 'packed':
            order.status = 'Packed'
            order.save()    
        elif new_status == 'shipped':
            order.status = 'Shipped'
            order.save()    
        elif new_status == 'delivered':
            order.status = 'Delevered'
            order.save()
            
    return redirect('custom_admin:order_management')   

def order_details(request, order_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    order = OrderDetails.objects.get(pk = order_id)
    order_address = order.address

    order_data = []
    user_name = order.user.fullname
    payment = order.payment  
    product_name = ", ".join(item.product.name for item in OrderItems.objects.filter(order_id=order))
    order_date = order.created_at.strftime("%d %b %Y")
    arriving_date = order.created_at + timedelta(days=5)
    status = order.status

    order_data.append({
        'order_id': order.id,
        'payment' : payment , 
        'user_name': user_name,
        'product_name': product_name,
        'order_date': order_date,
        'total': order.total,
        'arriving_date': arriving_date,
        'status': status,
    })
    
    if status == 'Delevered':
        step=[i for i in range(6)]
    elif status == 'Shipped':
        step=[i for i in range(5)]
    elif status == 'Packed':
        step=[i for i in range(4)]
    elif status == 'Accepted':
        step=[i for i in range(3)]
    elif status == 'Placed':
        step=[i for i in range(2)]
    else:
        step=[i for i in range(1)]
    
    return render(request, 'admin_order_details.html', {'order_data': order_data ,'order_id':order_id , 'order_address': order_address , 'step':step})     
        
def add_variants(request, product_id):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    product = get_object_or_404(Product, id=product_id)
    SIZE_CHOICES = Variant.SIZE_CHOICES

    if request.method == 'POST':
        for size_choice in SIZE_CHOICES:
            size = size_choice[0]
            quantity_key = f'quantity_{size}'
            
            raw_quantity = request.POST.get(quantity_key, '')


            quantity = int(raw_quantity) if raw_quantity.isdigit() else 0


            variants = Variant.objects.filter(product=product, size=size)

            if variants.exists():
                variant = variants.first()
            else:
                variant = Variant(product=product, size=size)
            variant.quantity = quantity
            variant.save()
            
            total_quantity = Variant.objects.filter(product_id=product_id).aggregate(Sum('quantity'))['quantity__sum']
            product.quantity = total_quantity
            product.save()

        return redirect('custom_admin:product_list')

    return render(request, 'add_variants.html', { 'SIZE_CHOICES': SIZE_CHOICES, 'product': product})

def generate_pdf(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    orders = OrderDetails.objects.all()
    context = {'orders': []}

    for order in orders:
        order_details = {
            'order': order,
            'order_items': OrderItems.objects.filter(order_id=order),
        }
        context['orders'].append(order_details)

    template_path = 'sales_report.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename="sales_report.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    # Create a PDF file
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

def generate_excel(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    orders = OrderDetails.objects.all()
    context = {'orders': []}

    for order in orders:
        order_details = {
            'order': order,
            'order_items': OrderItems.objects.filter(order_id=order),
        }
        context['orders'].append(order_details)
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'

    # Create an Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write header row
    header = ['Order ID', 'User Name', 'Product Name', 'Order Date', 'Arriving Date', 'Status']
    for col_num, value in enumerate(header, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet.cell(row=1, column=col_num, value=value)

    # Write data rows
    for row_num, order_details in enumerate(context['orders'], 1):
        order = order_details['order']
        worksheet.cell(row=row_num, column=1, value='GR000'+ str(order.id))
        worksheet.cell(row=row_num, column=2, value=order.user.fullname)
        
        for order_item_num, order_item in enumerate(order_details['order_items'], 1):
            worksheet.cell(row=row_num + order_item_num, column=3, value=order_item.product.name)
            # Add more columns for order item details
        order_date_str = order.created_at.strftime('%Y-%m-%d %H:%M:%S')
        worksheet.cell(row=row_num, column=4, value=order_date_str)
        if order.arriving_date:
            arriving_date_str = order.arriving_date.strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=row_num, column=5, value=arriving_date_str)
        worksheet.cell(row=row_num, column=6, value=order.status)
    # Save the workbook
    workbook.save(response)
    return response


def sales_report(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    return render(request, 'sales_report.html')


class OfferAdminView(View):
    template_name = 'offer_admin.html'

    def get(self, request):
        offers = Offer.objects.all()
        form = OfferForm()
        return render(request, self.template_name, {'offers': offers, 'form': form})

    def post(self, request):
        form = OfferForm(request.POST)

        if form.is_valid():
            name = form.cleaned_data['name']
            description = form.cleaned_data['description']
            discount_amount = form.cleaned_data['discount_amount']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            category = form.cleaned_data['category']

            # Create a new offer
            Offer.objects.create(
                name=name,
                description=description,
                discount_amount=discount_amount,
                start_date=start_date,
                end_date=end_date,
                category=category
            )

            return redirect('custom_admin:offer_admin')
        else:
            error_message = "Form is not valid"
            offers = Offer.objects.all()
            return render(request, self.template_name, {'offers': offers,'error_message':error_message ,'error_flag':True , 'form': form})
            

        # If the form is not valid, render the form with errors
        offers = Offer.objects.all()
        return render(request, self.template_name, {'offers': offers, 'form': form})

    
    
def delete_offer(request):
    
    if not request.user.is_staff:
        return HttpResponseForbidden("You do not have permission to access this page.")
    
    if request.method == 'GET':
        offer_id = request.GET.get('offer_id')
        if offer_id:
            # Retrieve the offer to be deleted
            offer = get_object_or_404(Offer, id=offer_id)
            
            # Delete the offer
            offer.delete()

            # Redirect to the offer admin page or another appropriate page
            return redirect('custom_admin:offer_admin')

    # Handle the case when offer_id is not provided or the request method is not GET
    return redirect('custom_admin:offer_admin')

def logout_view(request):
    return custom_logout